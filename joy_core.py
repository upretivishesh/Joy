import os
import re
import logging
import tempfile

import pandas as pd
import pdfplumber
from docx import Document
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from dateutil import parser as dt_parser
from dateutil.relativedelta import relativedelta

# ---------- QUIET LOGGING ----------
logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfplumber").setLevel(logging.ERROR)
logging.getLogger("PIL").setLevel(logging.ERROR)

# ---------- FILE READING HELPERS ----------

def read_pdf_fp(file_obj):
    """Read PDF text from a file-like object (UploadedFile / BytesIO)."""
    file_obj.seek(0)
    with pdfplumber.open(file_obj) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)

def read_docx_fp(file_obj):
    """Read DOCX text from a file-like object."""
    file_obj.seek(0)
    doc = Document(file_obj)
    return "\n".join(p.text for p in doc.paragraphs)

def read_any_fp(uploaded_file):
    """
    Detect extension from uploaded_file.name and read content accordingly.
    Works with Streamlit's UploadedFile or any object with .name and .read().
    """
    name = uploaded_file.name
    ext = os.path.splitext(name)[1].lower()
    uploaded_file.seek(0)

    if ext == ".pdf":
        return read_pdf_fp(uploaded_file)
    elif ext == ".docx":
        return read_docx_fp(uploaded_file)
    else:
        return uploaded_file.read().decode("utf-8", errors="ignore")

# ---------- RESUME PARSING ----------

def extract_details(text):
    """Extract name, mobile, email, location, total experience (string) + years (float)."""
    email = "-"
    mobile = "-"
    experience = "-"
    location = "-"
    name = "-"

    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # Email
    m = re.search(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", text)
    if m:
        email = m.group(0)
        email_local = email.split("@")[0]
    else:
        email_local = ""

    # Mobile
    m = re.search(r"(?:\+91[-\s]?)?\b\d{10}\b", text)
    if m:
        mobile = m.group(0)

    # Name – first all‑caps line near top
    for l in lines[:8]:
        if re.match(r"^[A-Z][A-Z ]+$", l) and "SOURCED" not in l:
            name = l.title()
            break
    if name == "-" and email_local:
        user = re.sub(r"\d+", "", email_local)
        user = re.sub(r"(vishesh)(upreti)", r"\1 \2", user, flags=re.IGNORECASE)
        parts = re.split(r"[._\-\s]+", user)
        parts = [p.capitalize() for p in parts if len(p) > 1]
        name = " ".join(parts) if parts else email

    # Location – "<City>, India"
    for l in lines[:15]:
        if "india" in l.lower() and "," in l:
            before = l.split(",")[0]
            city = before.split()[-1]
            location = city.title()
            break

    # Experience
    exp_periods = []
    now = dt_parser.parse("Nov 2025")

    for l in lines:
        for start_str, end_str in re.findall(
            r"([A-Za-z]+\s+\d{4})\s*[-–]\s*(Present|[A-Za-z]+\s+\d{4})", l
        ):
            try:
                start_date = dt_parser.parse(start_str)
                end_date = now if "present" in end_str.lower() else dt_parser.parse(end_str)
                rd = relativedelta(end_date, start_date)
                yrs = rd.years + rd.months / 12
                if yrs > 0:
                    exp_periods.append(yrs)
            except Exception:
                pass

    for sy, ey in re.findall(r"(\d{4})\s*[-–]\s*(\d{4})", text):
        try:
            yrs = int(ey) - int(sy)
            if yrs > 0:
                exp_periods.append(float(yrs))
        except Exception:
            pass

    total = sum(exp_periods)
    experience = f"{total:.1f} Years" if total > 0 else "-"
    years_float = total

    return {
        "Applicant Name": name or "-",
        "Mobile No.": mobile or "-",
        "Email Address": email or "-",
        "Location": location or "-",
        "Total Experience": experience or "-",
    }, years_float

# ---------- JD-DRIVEN FIT ANALYSIS + WHATSAPP ----------

def extract_jd_keywords(jd_text, top_n=30):
    """
    Simple keyword extractor: lowercase, split on non-letters,
    drop common stopwords, keep top-N most frequent words.
    """
    if not jd_text:
        return []

    text = jd_text.lower()
    tokens = re.split(r"[^a-z]+", text)
    stopwords = {
        "and", "or", "the", "a", "an", "of", "to", "for", "in", "on", "with",
        "is", "are", "as", "at", "by", "from", "this", "that", "will", "be",
        "you", "your", "our", "we", "they", "them", "their", "job", "role",
        "responsibilities", "requirements", "skills", "experience", "minimum",
        "must", "should", "have"
    }
    words = [w for w in tokens if len(w) > 2 and w not in stopwords]

    if not words:
        return []

    from collections import Counter
    counts = Counter(words)
    return [w for w, _ in counts.most_common(top_n)]

def analyze_fit(resume_text, years_float, jd_available: bool, jd_keywords=None):
    """
    JD-driven generic scoring:
    - Use overlap between resume_text and jd_keywords.
    - Use years of experience as additional weight.
    """
    if not jd_available or not jd_keywords:
        return "No JD-based scoring available"

    t = resume_text.lower()
    tokens = set(re.split(r"[^a-z]+", t))

    # Keyword overlap score
    matches = [kw for kw in jd_keywords if kw in tokens]
    kw_score = len(matches)

    # Experience weight
    if years_float >= 5:
        exp_score = 3
    elif years_float >= 3:
        exp_score = 2
    elif years_float >= 1:
        exp_score = 1
    else:
        exp_score = 0

    total_score = kw_score + exp_score

    if total_score >= 15:
        tag = "Strong fit"
    elif total_score >= 8:
        tag = "Partial fit"
    else:
        tag = "Not a fit"

    brief_bits = []
    if years_float:
        brief_bits.append(f"{years_float:.1f} yrs exp")
    if matches:
        brief_bits.append(f"{len(matches)} JD keyword matches")

    brief = ", ".join(brief_bits) if brief_bits else "limited match"
    return f"{tag} – {brief}"

def build_whatsapp_text(name, role):
    first_name = name.split()[0] if name and name != "-" else "there"
    return (
        f"Hi {first_name}, this is Joy from Seven Hiring.\n\n"
        f"I’m reaching out regarding an opportunity for {role}. "
        f"Your background looks relevant, and I’d like to check your interest and availability.\n\n"
        f"Could you please reply with:\n"
        f"1. Your Current CTC (fixed + variable)\n"
        f"2. Your Expected CTC\n"
        f"3. Your Notice Period\n"
        f"4. Any other interview processes or offers you’re currently involved in\n"
        f"5. Your virtual interview availability over the next 2–3 days\n\n"
        f"Once you share these, I’ll coordinate the next steps and schedule discussions accordingly."
    )

# ---------- MAIN API FOR app.py ----------

def process_resumes_with_jd(uploaded_jd, uploaded_resumes, role="Assistant Manager – Logistics"):
    """
    Main API used by app.py.
    uploaded_jd: single UploadedFile or None
    uploaded_resumes: list of UploadedFile
    Returns: (df, excel_bytes)
    """
    if uploaded_jd is not None:
        jd_text = read_any_fp(uploaded_jd)
        jd_available = bool(jd_text.strip())
        jd_keywords = extract_jd_keywords(jd_text) if jd_available else []
    else:
        jd_available = False
        jd_keywords = []

    rows = []
    for uf in uploaded_resumes:
        text = read_any_fp(uf)
        details, years_float = extract_details(text)
        remark = analyze_fit(
            text,
            years_float,
            jd_available,
            jd_keywords=jd_keywords,
        )
        details["Remark"] = remark
        details["WhatsApp Text"] = build_whatsapp_text(details["Applicant Name"], role)
        rows.append(details)

    if not rows:
        df = pd.DataFrame(columns=[
            "Applicant Name", "Mobile No.", "Email Address",
            "Location", "Total Experience", "Remark", "WhatsApp Text"
        ])
    else:
        df = pd.DataFrame(rows)

    # Build Excel in memory
    wb = Workbook()
    ws = wb.active
    ws.title = "Candidates"

    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)

    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    for cell in ws[1]:
        cell.font = bold
        cell.fill = yellow
        cell.alignment = center

    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 26.3
        for cell in col:
            cell.alignment = center

    remark_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Remark":
            remark_col_idx = idx
            break

    if remark_col_idx is not None:
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        orange = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            remark_cell = ws.cell(row=row, column=remark_col_idx)
            val = str(remark_cell.value or "").lower()
            if val.startswith("strong fit"):
                fill = green
            elif val.startswith("partial fit"):
                fill = orange
            elif val.startswith("not a fit"):
                fill = red
            else:
                fill = None
            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(tmp.name)
    tmp.seek(0)
    excel_bytes = tmp.read()
    tmp.close()
    os.remove(tmp.name)

    return df, excel_bytes
