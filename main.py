import os
import re
import logging

import pandas as pd
import pdfplumber
from docx import Document
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from dateutil import parser as dt_parser
from dateutil.relativedelta import relativedelta

# ---------- QUIET LOGGING ----------
logging.getLogger("pdfminer").setLevel(logging.ERROR)
logging.getLogger("pdfplumber").setLevel(logging.ERROR)
logging.getLogger("PIL").setLevel(logging.ERROR)

# ---------- CONFIG ----------
base_dir = r"C:\Users\vishe\OneDrive\Documents\Projects\Project_Joy"
resume_folder = os.path.join(base_dir, "resumes")
jd_folder = os.path.join(base_dir, "jd")
output_file = os.path.join(base_dir, "outputs", "candidates_details.xlsx")

columns = [
    "Applicant Name", "Mobile No.", "Email Address",
    "Location", "Total Experience", "Remark", "WhatsApp Text"
]

# ---------- BASIC IO ----------
def read_pdf(path):
    with pdfplumber.open(path) as pdf:
        return "\n".join(page.extract_text() or "" for page in pdf.pages)

def read_docx(path):
    doc = Document(path)
    return "\n".join(p.text for p in doc.paragraphs)

def read_any(path):
    ext = os.path.splitext(path)[1].lower()
    if ext == ".pdf":
        return read_pdf(path)
    elif ext == ".docx":
        return read_docx(path)
    else:
        with open(path, "r", encoding="utf-8", errors="ignore") as f:
            return f.read()

def load_jd_text_and_flag():
    if not os.path.isdir(jd_folder):
        return "", False
    for fname in os.listdir(jd_folder):
        path = os.path.join(jd_folder, fname)
        if os.path.isfile(path) and path.lower().endswith((".pdf", ".docx", ".txt")):
            return read_any(path), True
    return "", False

JD_TEXT, JD_AVAILABLE = load_jd_text_and_flag()

# ---------- RESUME PARSING ----------
def extract_details(text):
    email = "-"
    mobile = "-"
    experience = "-"
    location = "-"
    name = "-"

    lines = [l.strip() for l in text.splitlines() if l.strip()]

    # Email
    m = re.search(r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}\b", text)
    if m:
        email = m.group(0)
        email_local = email.split("@")[0]
    else:
        email_local = ""

    # Mobile
    m = re.search(r"(?:\+91[-\s]?)?\b\d{10}\b", text)
    if m:
        mobile = m.group(0)

    # Name – first all‑caps line near top (e.g. "MEENAKSHI NEGI")
    for l in lines[:8]:
        if re.match(r"^[A-Z][A-Z ]+$", l) and "SOURCED" not in l:
            name = l.title()
            break
    if name == "-" and email_local:
        user = re.sub(r"\d+", "", email_local)
        # small custom split for your own case
        user = re.sub(r"(vishesh)(upreti)", r"\1 \2", user, flags=re.IGNORECASE)
        parts = re.split(r"[._\-\s]+", user)
        parts = [p.capitalize() for p in parts if len(p) > 1]
        name = " ".join(parts) if parts else email

    # Location – "<City>, India" near top
    for l in lines[:15]:
        if "india" in l.lower() and "," in l:
            before = l.split(",")[0]
            city = before.split()[-1]
            location = city.title()
            break

    # Experience – month‑year ranges & year–year ranges
    exp_periods = []
    now = dt_parser.parse("Nov 2025")

    for l in lines:
        for start_str, end_str in re.findall(
            r"([A-Za-z]+\s+\d{4})\s*[-–]\s*(Present|[A-Za-z]+\s+\d{4})", l
        ):
            try:
                start_date = dt_parser.parse(start_str)
                end_date = now if "present" in end_str.lower() else dt_parser.parse(end_str)
                rd = relativedelta(end_date, start_date)
                yrs = rd.years + rd.months / 12
                if yrs > 0:
                    exp_periods.append(yrs)
            except Exception:
                pass

    for sy, ey in re.findall(r"(\d{4})\s*[-–]\s*(\d{4})", text):
        try:
            yrs = int(ey) - int(sy)
            if yrs > 0:
                exp_periods.append(float(yrs))
        except Exception:
            pass

    total = sum(exp_periods)
    experience = f"{total:.1f} Years" if total > 0 else "-"
    years_float = total

    return {
        "Applicant Name": name or "-",
        "Mobile No.": mobile or "-",
        "Email Address": email or "-",
        "Location": location or "-",
        "Total Experience": experience or "-",
    }, years_float

# ---------- FIT ANALYSIS ----------
def analyze_fit(resume_text, years_float, location):
    if not JD_AVAILABLE:
        return "No JD file found"

    t = resume_text.lower()

    in_logistics = any(w in t for w in ["logistics", "supply chain", "warehouse", "transport"])
    has_export = any(w in t for w in ["export", "import", "cross border", "customs", "incoterm", "hs code"])
    has_docs = any(w in t for w in [
        "documentation", "export documentation", "shipping documents", "bill of lading", "bl "
    ])
    has_modes = any(w in t for w in ["road", "sea", "air"])
    in_blr = "bengaluru" in t or "bangalore" in t or location.lower() in ["bengaluru", "bangalore"]

    score = 0
    score += 2 if years_float >= 3 else 0
    score += 2 if in_logistics else 0
    score += 2 if has_export else 0
    score += 1 if has_docs else 0
    score += 1 if has_modes else 0
    score += 1 if in_blr else 0

    if score >= 6:
        tag = "Strong fit"
    elif score >= 4:
        tag = "Partial fit"
    else:
        tag = "Not a fit"

    bullets = []
    if years_float:
        bullets.append(f"{years_float:.1f} yrs exp")
    if in_logistics:
        bullets.append("logistics/supply chain")
    if has_export:
        bullets.append("import/export")
    if has_docs:
        bullets.append("docs")
    if has_modes:
        bullets.append("road/sea/air")
    if in_blr:
        bullets.append("BLR-based")

    brief = ", ".join(bullets[:4]) if bullets else "limited match"
    return f"{tag} – {brief}"

# ---------- WHATSAPP TEXT ----------
def build_whatsapp_text(name):
    first_name = name.split()[0] if name and name != "-" else "there"
    return (
        f"Hi {first_name}, this is Joy from Seven Hiring.\n\n"
        f"I’m reaching out regarding an opportunity for Assistant Manager – Logistics with Atomgrid in Bengaluru. "
        f"Your background in logistics and supply chain looks relevant, and I’d like to check your interest and availability.\n\n"
        f"Could you please reply with:\n"
        f"1. Your Current CTC (fixed + variable)\n"
        f"2. Your Expected CTC\n"
        f"3. Your Notice Period\n"
        f"4. Any other interview processes or offers you’re currently involved in\n"
        f"5. Your virtual interview availability over the next 2–3 days\n\n"
        f"Once you share these, I’ll coordinate the next steps and schedule discussions accordingly."
    )

# ---------- EXCEL FORMAT WITH FULL-ROW COLORS ----------
def format_excel(path):
    wb = load_workbook(path)
    ws = wb.active
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")

    # Header styling
    for cell in ws[1]:
        cell.font = bold
        cell.fill = yellow
        cell.alignment = center

    # Column widths + center alignment
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 26.3
        for cell in col:
            cell.alignment = center

    # Find Remark column index
    remark_col_idx = None
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value == "Remark":
            remark_col_idx = idx
            break

    if remark_col_idx is not None:
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        orange = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
        red = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

        for row in range(2, ws.max_row + 1):
            remark_cell = ws.cell(row=row, column=remark_col_idx)
            val = str(remark_cell.value or "").lower()
            if val.startswith("strong fit"):
                fill = green
            elif val.startswith("partial fit"):
                fill = orange
            elif val.startswith("not a fit"):
                fill = red
            else:
                fill = None

            if fill:
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=row, column=col).fill = fill

    wb.save(path)

# ---------- MAIN FLOW ----------
os.makedirs(os.path.join(base_dir, "outputs"), exist_ok=True)
rows = []

for fname in os.listdir(resume_folder):
    path = os.path.join(resume_folder, fname)
    if os.path.isfile(path) and path.lower().endswith((".pdf", ".docx")):
        print("Parsing:", fname)
        text = read_any(path)
        details, years_float = extract_details(text)
        remark = analyze_fit(text, years_float, details["Location"])
        details["Remark"] = remark
        details["WhatsApp Text"] = build_whatsapp_text(details["Applicant Name"])
        rows.append(details)

if not rows:
    print("No applicant data found.")
else:
    df_new = pd.DataFrame(rows, columns=columns)
    if os.path.exists(output_file):
        existing = pd.read_excel(output_file)
        df = pd.concat([existing, df_new], ignore_index=True)
    else:
        df = df_new
    df.to_excel(output_file, index=False)
    format_excel(output_file)
    print("Done. Written to:", output_file)
